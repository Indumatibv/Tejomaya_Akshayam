# conda activate tejomaya 
# python -m agents.Parsing_agent

#!/usr/bin/env python
# agents/parsing_agent.py
# ============================================================
# REGULATIONS-ONLY PARSING & SUMMARY (TEJOMAYA v1)
# ============================================================
#!/usr/bin/env python

import sys
from pathlib import Path

# ✅ FORCE project root into PYTHONPATH
BASE_DIR = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(BASE_DIR))

from storage.minio_client import MinIOClient

import os
import re
import logging
from pathlib import Path
import pandas as pd
from dotenv import load_dotenv
from unstructured.partition.pdf import partition_pdf
from openpyxl import load_workbook, Workbook
from datetime import datetime
import torch
import warnings
import time

# from langchain.llms import Ollama
from langchain_community.llms import Ollama

#from langchain_ollama import OllamaLLM

from langchain.prompts import PromptTemplate

# ---------------------- Logging ----------------------
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[logging.StreamHandler()]
)

warnings.filterwarnings("ignore")
load_dotenv()

# ---------------------- GPU Detection ----------------------
device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
if device.type == "cuda":
    os.environ["OLLAMA_USE_GPU"] = "1"
    os.environ["OLLAMA_NUM_GPU_LAYERS"] = "35"
    print(f"Using GPU: {torch.cuda.get_device_name(0)}")
else:
    os.environ["OLLAMA_USE_GPU"] = "0"  
    print("Using CPU")  

 # 👇 ADD THIS FUNCTION HERE (after GPU detection block)
# def print_device_usage(operation: str):
#     device_name = torch.cuda.get_device_name(0) if torch.cuda.is_available() else "CPU"
#     gpu_layers = os.environ.get("OLLAMA_NUM_GPU_LAYERS", "0")
#     print(f"🔧 [{operation}] Using {device_name} (GPU layers: {gpu_layers})")
# ---------------------- LLM ----------------------
llm = Ollama(model="mistral:latest")

# ============================================================
CREATED_EXCELS = set()

# ============================================================
# REGULATIONS SUMMARY PROMPT (FINAL, AUTHORITATIVE)
# ============================================================

REGULATIONS_PROMPT = PromptTemplate(
    template="""
You are a senior regulatory analyst preparing a concise, client-ready summary of a SEBI regulation document
for business stakeholders and compliance teams.

This document sets out an entire regulatory framework with detailed legal provisions.
The reader will NOT open the original regulation, so the summary must explain in simple words what it covers and why it matters.

CONTENT FOCUS:
- Explain at a high level what the regulation governs (e.g., REITs, InvITs, intermediaries, market participants)
- State clearly who it applies to (e.g., sponsors, managers, trustees, listed entities, intermediaries)
- Highlight the key compliance and governance areas (e.g., registration, listing, valuation, disclosures, reporting, conduct requirements)
- Summarise the core responsibilities and obligations imposed on regulated entities
- Convey the overall regulatory intent and scope (e.g., transparency, investor protection, market integrity, better governance)

DO NOT:
- List clauses, chapters, or definitions
- Mention page counts, circular numbers, or legal citations
- Quote or paraphrase legal provisions verbatim
- Go into historical amendments or minor editorial changes

STYLE AND LENGTH:
- Use simple, non-technical language that a business reader can understand
- Keep the summary short but sufficient to give a clear picture of the framework
- Limit strictly to 5–6 sentences
- Write as a single coherent paragraph

STARTING RULE (MANDATORY):
- The first sentence MUST start with:
  “This SEBI regulation sets out the regulatory framework for …”

Text:
{text}

Final Summary (entire answer MUST be inside double quotes):
""",
    input_variables=["text"]
)

# ============================================================

AMENDMENT_REGULATIONS_PROMPT = PromptTemplate(
    template="""
You are a senior regulatory analyst preparing a client-ready summary of a SEBI amendment to existing regulations.

The reader is a business or compliance professional who will NOT read the original document.
The summary must clearly explain in simple language what has changed and why it matters in practice.

STRUCTURE (MANDATORY):
- Output ONLY bullet points (no paragraph introduction)
- Use ONLY black bullet points “•”
- The FIRST bullet MUST start with: “SEBI has amended or added a regulation to …”
- The FIRST bullet must be ONE sentence that gives a high-level overview of the main changes
- ALL remaining bullets must each describe ONE important new or amended provision and its practical effect
- Do NOT use nested bullets or sub-points; every change must be its own “•” bullet

CONTENT RULES (GENERIC):
- Focus on changes that materially affect:
  • who is covered or brought into scope, or
  • what conditions, qualifications, certifications, limits or thresholds apply, or
  • what information, processes, disclosures or infrastructure are now required, or
  • how compliance, governance or investor protection will work in practice.
- Each bullet must describe the real-world outcome or impact, not the drafting mechanics
- At least ONE bullet must clearly state the impact or benefit for the public, investors or other relevant stakeholders
- Ignore minor editorial or housekeeping changes (like word substitutions, formatting, removal of fax numbers) unless they meaningfully change compliance or process
- Do NOT use phrases like:
  “definition of”, “the term”, “has been defined”, “has been introduced”
- Do NOT use colon-style labels or headings at the start of any bullet (e.g., “Impact:”, “Change:”, “Key update:”)
- Do NOT quote or paraphrase legal definitions verbatim
- Do NOT mention clause numbers, insertions, omissions, substitutions, schedules, or form item numbers; summarise their practical effect instead

FORMAT RULES (STRICT):
- Write 4 or 5 bullet points in total
- Each bullet must be ONE concise sentence
- Do NOT use numbering (1., 2., 3.) or dashes (-)
- No bullet may end with a colon

TONE:
- Plain, professional, and client-facing
- Written like a regulatory update for senior stakeholders
- Easy to scan and understand for a non-legal audience

Text:
{text}

Final Summary (use ONLY black bullet points “•”):
""",
    input_variables=["text"]
)

# ============================================================
# CIRCULARS SUMMARY PROMPT
# ============================================================
CIRCULARS_PROMPT = PromptTemplate(
    template="""
You are a regulatory analyst writing a short, client-ready summary of a SEBI circular.

SEBI circulars provide clarifications, implementation guidance, amendments, or new requirements.
The reader will NOT open the original circular, so the summary must clearly explain what has changed and what the reader needs to know in practice.

CONTENT RULES:
- Clearly state the main clarification, change, amendment, or new requirement introduced by the circular in plain language.
- Explicitly mention what is NEW or DIFFERENT compared to earlier practice or requirements.
- Include any key conditions, thresholds, timelines, applicability, or affected entities only if they matter for compliance.
- Briefly include other important informational points mentioned in the circular if they help the reader understand obligations or impact.
- Do NOT use vague phrases such as “other details are specified in the circular or website”; instead, state those details concisely.
- Avoid background context, policy intent, or legal reasoning unless it directly affects what must be done.
- Do NOT include circular numbers, file references, email IDs, internal committees, venues, or legal citations.

FORMAT RULES:
- Output ONLY bullet points (no headings or paragraphs).
- Use ONLY black bullet points “•”.
- Write 2 to 4 bullet points in total.
- Each bullet must be ONE clear, concise sentence.
- Do NOT use numbering, sub-bullets, or multi-line bullets.

STARTING RULE (MANDATORY):
- The FIRST bullet MUST start with one of the following (choose what fits best):
  “SEBI has issued this circular and introduced …”
  “SEBI has issued this circular and clarified …”
  “SEBI has issued this circular and amended …”
  “SEBI has issued this circular and changed …”

TONE:
- Plain, client-facing language.
- Simple, direct, and easy to understand for non-legal readers.
- Avoid technical or regulatory jargon wherever possible.

Text:
{text}

Final Summary:
""",
    input_variables=["text"]
)

# CIRCULARS_PROMPT = PromptTemplate(
#     template="""
# You are a regulatory analyst writing a short, client-ready summary of a SEBI circular.

# SEBI circulars provide clarifications, implementation guidance, or changes to existing requirements.
# The reader will NOT open the original circular, so the summary must tell them clearly what has changed and what they need to know in practice.

# CONTENT RULES:
# - Clearly state the main clarification, change, or requirement introduced by the circular in simple language.
# - Mention any key conditions, thresholds, timelines, or applicability (for example, which entities or transactions are covered) only if they are important for compliance.
# - Briefly mention other important informational points instead of using vague phrases like “other details are specified in the circular”.
# - Avoid technical or legal jargon and avoid background or policy rationale unless it directly affects what must be done.
# - Do NOT include circular numbers, SEBI file numbers, internal processes, committee names, venue details, or long legal citations.

# FORMAT RULES:
# - Output ONLY bullet points (no paragraphs or headings).
# - Use ONLY black bullet points “•”.
# - Write 2 to 4 bullet points in total.
# - Each bullet must be ONE clear, concise sentence.
# - Do NOT use numbering (1., 2., 3.) or sub-bullets.

# STARTING RULE (MANDATORY):
# - The FIRST bullet MUST start with:
#   “SEBI has issued this circular and …”

# TONE:
# - Plain language and client-facing.
# - Crisp, direct, and easy to understand for non-legal readers.

# Text:
# {text}

# Final Summary:
# """,
#     input_variables=["text"]
# )

# ============================================================
# PRESS RELEASES SUMMARY PROMPT
# ============================================================

PRESS_RELEASE_PROMPT = PromptTemplate(
    template="""
You are preparing a short, client-ready summary of a SEBI press release.

SUB-DOMAIN: Press Releases

ABOUT:
Press releases are informational and communicate key decisions, actions, warnings, clarifications, or outcomes.
Summarise ONLY the primary outcomes that the reader must know.

STRICT RULES:
- Output ONLY bullet points
- Use ONLY black bullet points “•”
- Write ONLY what SEBI has decided, clarified, approved, warned, or stated
- Do NOT include background, explanations, names, dates, venues, links, or references to media reports
- If there are multiple key outcomes, write 2–3 bullets (no more)

STARTING RULE (MANDATORY):
- Each bullet MUST start with:
  “The press release issued states that …”

STYLE:
- One sentence per bullet
- Plain, direct, client-facing language
- No legal or procedural wording

Text:
{text}

Final Summary:
""",
    input_variables=["text"]
)

# ============================================================
# CONSULTATION PAPERS SUMMARY PROMPT
# ============================================================

# CONSULTATION_PAPER_PROMPT = PromptTemplate(
#     template="""
# You are a regulatory analyst preparing a client-ready summary of a SEBI consultation paper.

# Consultation papers propose concrete regulatory changes and seek public feedback.
# The reader will NOT read the original document.

# ABSOLUTE GOAL:
# - After reading the summary, the reader must clearly understand EXACTLY what regulatory areas are proposed to be changed and the objective of the consultation paper, without opening the document.

# CRITICAL READING INSTRUCTION (MANDATORY):
# - The document may describe proposed changes inside tables (for example, columns such as “Current Provision”, “Proposed Change”, “Rationale”).
# - You MUST read and interpret these tables.
# - You MUST summarise the substance of the proposed changes shown in tables.
# - Do NOT repeat section headings such as “certain provisions”; always expand them using the actual table content.

# STRICT FORMAT RULES (MANDATORY):
# - Output ONLY black bullet points “•”
# - Do NOT write any paragraph text
# - Do NOT use hyphens (-), sub-bullets, or nested points
# - Write EXACTLY 3 or 4 bullet points
# - Each bullet must be ONE complete sentence

# STRUCTURE (MANDATORY):
# - The FIRST bullet MUST start exactly with:
#   “SEBI has issued this consultation paper proposing the following changes …”
# - The NEXT bullets must each describe ONE specific proposed regulatory change
# - The FINAL bullet MUST state that SEBI is seeking public comments, views, or suggestions

# MANDATORY SENTENCE PATTERN FOR PROPOSED CHANGES (CRITICAL):
# - Every proposed-change bullet MUST follow this structure:
#   “Proposing changes to <specific regulatory area> to <specific nature of the change>.”
# - The <specific regulatory area> MUST be explicitly named, such as:
#   issuance of securities, registration and transfer of securities,
#   disclosure requirements, post-issue compliance,
#   operational and record-keeping requirements under Schedule VII,
#   governance obligations, scope of applicability.
# - Bullets that do NOT clearly name the regulatory area are INVALID and must be rewritten.

# SPECIFICITY RULE (NON-NEGOTIABLE):
# - Each proposed-change bullet MUST state BOTH:
#   • what regulatory area is affected, AND
#   • how that area is being changed, based on the document text or tables.
# - Do NOT describe proposals at an abstract, intent-based, or heading-based level.

# HARD PROHIBITIONS (STRICTLY ENFORCED):
# - Do NOT use vague or placeholder phrases such as:
#   “certain provisions”, “various changes”, “other measures”,
#   “related aspects”, “market developments”, or “regulatory landscape”.
# - Do NOT copy consultation questions or section titles as summary points.
# - Do NOT frame proposals as questions.
# - Do NOT explain why the change is proposed.
# - Do NOT mention consultation timelines, dates, emails, links,
#   clause numbers, or legal drafting language.

# PUBLIC COMMENTS RULE (MANDATORY):
# - The public comments bullet MUST be exactly:
#   “SEBI is seeking public comments, views, or suggestions on the proposed changes.”
# - Do NOT include, paraphrase, or list consultation questions.
# - Do NOT include dates, deadlines, links, URLs, or submission instructions.

# QUALITY BAR (MANDATORY SELF-CHECK):
# - If a sentence could make the reader ask “what exactly is changing?”, it is INVALID
#   and MUST be rewritten with concrete detail taken from the document or tables.


# Text:
# {text}

# Final Summary (use ONLY black bullet points “•”):
# """,
#     input_variables=["text"]
# )
CONSULTATION_PAPER_PROMPT = PromptTemplate(
    template="""
You are a regulatory analyst preparing a client-ready summary of a SEBI consultation paper.

Consultation papers propose concrete regulatory changes and seek public feedback.
The reader will NOT read the original document.

ABSOLUTE GOAL:
- After reading the summary, the reader must clearly understand EXACTLY what regulatory areas are proposed to be changed and the objective of the consultation, without opening the document.

CRITICAL READING INSTRUCTION (MANDATORY):
- The document may describe proposed changes inside tables (for example, columns such as “Current Provision”, “Proposed Change”, “Rationale”).
- You MUST read and interpret these tables.
- You MUST summarise the substance of the proposed changes shown in tables.
- Do NOT repeat section headings such as “certain provisions”; always expand them using the actual table content.

STRICT FORMAT RULES (MANDATORY):
- Output ONLY black bullet points “•”
- Do NOT write any paragraph text
- Do NOT use hyphens (-), sub-bullets, or nested points
- Write EXACTLY 4 bullet points
- Each bullet must be ONE complete sentence

STRUCTURE (MANDATORY):
- The FIRST bullet MUST start exactly with:
  “SEBI has issued this consultation paper proposing the following changes …”
- The SECOND bullet MUST clearly state the OBJECTIVE of the consultation paper,
  explicitly describing what SEBI aims to achieve (for example, improving trading
  processes at stock exchanges, strengthening compliance, enhancing transparency, etc.)
- The THIRD bullet MUST describe ONE specific proposed regulatory change
- The FOURTH (FINAL) bullet MUST state that SEBI is seeking public comments AND
  MUST include the deadline for submitting comments, if mentioned in the document

MANDATORY SENTENCE PATTERN FOR PROPOSED CHANGES (CRITICAL):
- Every proposed-change bullet MUST follow this structure:
  “Proposing changes to <specific regulatory area> to <specific nature of the change>.”
- The <specific regulatory area> MUST be explicitly named, such as:
  issuance of securities, registration and transfer of securities,
  disclosure requirements, post-issue compliance,
  operational and record-keeping requirements under Schedule VII,
  governance obligations, scope of applicability.
- Bullets that do NOT clearly name the regulatory area are INVALID and must be rewritten.

OBJECTIVE EXTRACTION RULE (MANDATORY):
- The objective bullet MUST be derived from the document’s stated purpose,
  background, or explanatory sections.
- Do NOT restate headings such as “Objective of the Consultation Paper”.
- Do NOT use vague intent statements; the objective must be concrete and specific.

SPECIFICITY RULE (NON-NEGOTIABLE):
- Each proposed-change bullet MUST state BOTH:
  • what regulatory area is affected, AND
  • how that area is being changed, based on the document text or tables.
- Do NOT describe proposals at an abstract, intent-based, or heading-based level.

HARD PROHIBITIONS (STRICTLY ENFORCED):
- Do NOT use vague or placeholder phrases such as:
  “certain provisions”, “various changes”, “other measures”,
  “related aspects”, “market developments”, or “regulatory landscape”.
- Do NOT copy consultation questions or section titles as summary points.
- Do NOT frame proposals as questions.
- Do NOT explain why the change is proposed.
- Do NOT include emails, URLs, submission instructions, clause numbers,
  or legal drafting language.

PUBLIC COMMENTS & DEADLINE RULE (MANDATORY):
- The final bullet MUST:
  • state that SEBI is seeking public comments, views, or suggestions, AND
  • explicitly mention the deadline for submitting comments if provided.
- If the deadline is NOT mentioned in the document, state only that
  SEBI is seeking public comments without inventing a date.

QUALITY BAR (MANDATORY SELF-CHECK):
- If a sentence could make the reader ask “what exactly is changing or being achieved?”,
  it is INVALID and MUST be rewritten with concrete detail taken from the document or tables.

Text:
{text}

Final Summary (use ONLY black bullet points “•”):
""",
    input_variables=["text"]
)

# ============================================================
# MASTER CIRCULARS SUMMARY PROMPT
# ============================================================

MASTER_CIRCULAR_PROMPT = PromptTemplate(
    template="""
You are a regulatory analyst preparing a short, client-ready summary of a SEBI Master Circular.

Master Circulars consolidate all existing circulars and directions on a topic into a single updated reference document.

FORMAT RULES (MANDATORY):
- Output ONLY black bullet points “•”
- Write EXACTLY 3 bullet points
- Each bullet must be ONE sentence
- No sub-bullets, numbering, links, or references

CONTENT RULES:
- State that SEBI has issued a Master Circular
- Clearly mention the topic it covers
- Mention that it consolidates and supersedes earlier circulars where specified
- Emphasize that it is issued for ease of reference and is to be followed going forward
- Do NOT mention websites, URLs, departments, dates, annexures, or legal effects (such as rescission, savings, or prior actions)
- Do NOT describe structure, categories, or internal organisation
- Do NOT explain legal consequences, how prior actions are treated, or any procedural details

STARTING RULE (MANDATORY):
- The FIRST bullet MUST start exactly with:
  “SEBI has issued a master circular for …”

Text:
{text}

Final Summary:
""",
    input_variables=["text"]
)


# ============================================================
# NOTIFICATIONS SUMMARY PROMPT
# ============================================================

NOTIFICATIONS_PROMPT = PromptTemplate(
    template="""
You are a regulatory analyst preparing a short, client-ready summary of an IFSCA notification.

SUB-DOMAIN: Notifications

ABOUT:
IFSCA notifications are official announcements that introduce new rules, amendments, designations,
exemptions, or regulatory clarifications.

The reader will NOT open the original notification.
The summary must give a clear understanding of:
• what the notification does,
• why it was issued,
• when it applies, and
• who is impacted.

MANDATORY STARTING RULE:
- The summary MUST start exactly with:
  “This notification dated <issuance date> notifies …”
- If the exact issuance date is clearly available in the text, use it.
- If not clearly available, write:
  “This notification notifies …” (without a date).

CONTENT RULES (STRICT):
- Clearly state the key regulatory action taken (e.g., designation, amendment, inclusion, exemption, clarification).
- Briefly explain the purpose or regulatory intent only if it helps understanding.
- Explicitly mention:
  • applicability (who or what is covered), and
  • effective date (publication date or stated effective date), if mentioned.
- Focus on regulatory impact and compliance relevance.
- Summarise substance, NOT drafting mechanics.

DETAILS TO IGNORE (MUST NOT APPEAR):
- Internal circulation notes
- Annexure references or page numbers
- Signatures and sign-off blocks
- Email IDs, phone numbers, URLs
- File numbers, notification numbers, Gazette references
- Irrelevant numerical references

FORMAT RULES (MANDATORY):
- Output ONLY black bullet points “•”
- Write EXACTLY 2 or 3 bullet points
- Each bullet must be ONE clear sentence
- No headings, no numbering, no sub-bullets
- No quotes

STYLE:
- Plain, professional, client-facing
- No legal jargon
- Concise and scannable

Text:
{text}

Final Summary (use ONLY black bullet points “•”):
""",
    input_variables=["text"]
)

# ============================================================
# Quality Check
# ============================================================

# SUMMARY_CLEANER_PROMPT = PromptTemplate(
#     template="""
# You are reviewing a generated regulatory summary before final publication.

# TASK:
# - Remove any bullet points or sentences that are vague, generic, or non-informative.
# - Keep ONLY bullets that state concrete regulatory actions, obligations, scope, or outcomes.
# - Do NOT rewrite or invent new content.
# - Do NOT add missing details.
# - If a bullet contains vague phrases and no concrete regulatory substance, REMOVE it entirely.
# - If a bullet is concrete and clear, KEEP it unchanged.

# VAGUE PHRASES INCLUDE (NON-EXHAUSTIVE):
# - certain provisions
# - various changes
# - other measures
# - related aspects
# - market developments
# - regulatory landscape
# - streamlining processes
# - reviewing provisions
# - considering changes

# RULES:
# - Preserve the original bullet formatting.
# - Preserve the original order of remaining bullets.
# - Output ONLY the cleaned summary.
# - If all bullets are valid, return the summary unchanged.
# - If all bullets are vague, return "NA".

# Summary to review:
# {summary}

# Cleaned Summary:
# """,
#     input_variables=["summary"]
# )

SUMMARY_CLEANER_PROMPT = PromptTemplate(
    template="""
You are reviewing a generated regulatory summary before final publication.

TASK:
- Remove any bullet points or sentences that are vague, generic, or non-informative.
- Keep ONLY bullets that state concrete regulatory actions, obligations, scope, or outcomes.
- Do NOT rewrite or invent new content.
- Do NOT add missing details.
- If a bullet contains vague phrases and no concrete regulatory substance, REMOVE it entirely.
- If a bullet is concrete and clear, KEEP it unchanged.

VAGUE PHRASES INCLUDE (NON-EXHAUSTIVE):
- certain provisions
- various changes
- other measures
- related aspects
- market developments
- regulatory landscape
- streamlining processes
- reviewing provisions
- considering changes
- details are available on the website
- information can be accessed online
- uploaded details found on the website
- refer to the attached document
- further details are provided below
- further information is available
- Circulars can be accessed 
- Legal - > Circulars
- Legal - &gt; Circulars
- website under the link
- www.sebi.gov.in

ALSO REMOVE BULLETS THAT ONLY CONTAIN:
- venue, location, or event details
- protocol, inauguration, or ceremonial references
- names or presence of officials without any regulatory decision
- background or contextual information that does not affect the substance of the decision

RULES:
- Preserve the original bullet formatting.
- Preserve the original order of remaining bullets.
- Output ONLY the cleaned summary.
- If all bullets are valid, return the summary unchanged.
- If all bullets are vague, return "NA".

Summary to review:
{summary}

Cleaned Summary:
""",
    input_variables=["summary"]
)
# ============================================================
# PATHS
# ============================================================

BASE_DIR = Path(__file__).resolve().parent.parent
DATA_DIR = BASE_DIR / "data"
OUTPUT_EXCEL_DIR = DATA_DIR / "output_excels"
OUTPUT_EXCEL_DIR.mkdir(parents=True, exist_ok=True)

# ============================================================
# WEEKLY FOLDER
# ============================================================

def is_amendment_regulation(text: str) -> bool:
    return bool(re.search(r'\bamendment\b', text, re.IGNORECASE))

def get_week_folder():
    import json

    week_json = DATA_DIR / "week_range.json"
    if not week_json.exists():
        raise RuntimeError("week_range.json missing. Run searching_agent first.")

    with open(week_json, "r") as f:
        week = json.load(f)

    ws = datetime.strptime(week["week_start"], "%Y-%m-%d")
    we = datetime.strptime(week["week_end"], "%Y-%m-%d")

    folder = OUTPUT_EXCEL_DIR / f"{ws:%Y-%m-%d}_to_{we:%Y-%m-%d}"

    if folder.exists():
        import shutil
        shutil.rmtree(folder)

    folder.mkdir(parents=True)
    return folder


WEEK_FOLDER = get_week_folder()
logging.info(f"Weekly folder → {WEEK_FOLDER}")

# ============================================================

def clean_summary_with_llm(summary: str) -> str:
    if not summary or summary.strip() == "NA":
        return summary

    cleaned = llm.invoke(
        SUMMARY_CLEANER_PROMPT.format(summary=summary)
    ).strip()

    return cleaned or "NA"

# ============================================================
# PDF EXTRACTION (REGULATIONS-SAFE)
# ============================================================

def extract_pdf_text(pdf_path: str) -> str:
    raw = partition_pdf(
        filename=str(pdf_path),
        strategy="fast",
        include_page_breaks=False
    )

    text = "\n".join(str(el) for el in raw if el).strip()

    if not text:
        logging.info("Fallback to hi_res OCR")
        raw = partition_pdf(filename=str(pdf_path), strategy="hi_res")
        text = "\n".join(str(el) for el in raw if el).strip()

    return text
    

# ============================================================
# REGULATIONS TEXT FILTERING
# (THIS IS THE CORE LOGIC)
# ============================================================

def extract_regulation_core(text: str) -> str:
    lines = text.splitlines()
    keep = []
    capture = False

    for line in lines:
        clean = line.strip()

        if not clean:
            continue

        #  Start only at actual regulation body
        if re.search(r'in exercise of the powers conferred', clean, re.IGNORECASE):
            capture = True
            continue

        if not capture:
            continue

        #  Skip amendment history / compilation noise
        if re.search(
            r'(first|second|third|fourth|fifth|sixth|seventh|eighth)\s+amendment',
            clean,
            re.IGNORECASE
        ):
            continue

        if re.search(r'as amended upto|as amended up to', clean, re.IGNORECASE):
            continue

        #  Skip definition-only enumerations
        if re.match(r'^\([a-z]+\)', clean):
            continue

        #  Keep real regulatory substance
        if len(clean) > 40:
            keep.append(clean)

        if len(keep) >= 2500:
            break

    return "\n".join(keep)


# ============================================================
# SUMMARY GENERATION (NO CHUNKS)
# ============================================================

def generate_regulation_summary(text: str) -> str:
    core_text = extract_regulation_core(text)
    core_text = core_text[:12000]

    if is_amendment_regulation(core_text):
        prompt = AMENDMENT_REGULATIONS_PROMPT
    else:
        prompt = REGULATIONS_PROMPT

    summary = llm.invoke(
        prompt.format(text=core_text)
    ).strip()

    return summary or "NA", core_text

# ============================================================
# EXCEL UPDATE
# ============================================================

def update_excel(row: pd.Series):
    vertical = row["Verticals"]
    sub = row["SubCategory"]

    excel_path = WEEK_FOLDER / f"{vertical}.xlsx"

    CREATED_EXCELS.add(excel_path.name)  #  TRACK OWNERSHIP

    if excel_path.exists():
        wb = load_workbook(excel_path)
    else:
        wb = Workbook()
        wb.remove(wb.active)

    if sub not in wb.sheetnames:
        ws = wb.create_sheet(title=sub)
        ws.append(list(row.index))
    else:
        ws = wb[sub]

    ws.append([row.get(c, "NA") for c in row.index])
    wb.save(excel_path)
    wb.close()

# ============================================================
# PROCESS SINGLE PDF (REGULATIONS ONLY)
# ============================================================

def process_regulation_pdf(row: pd.Series):
    pdf_path = Path(row["Path"])

    try:
        text = extract_pdf_text(pdf_path)
        summary, embedding_text = generate_regulation_summary(text)

        # row["Summary"] = summary
        row["Summary"] = clean_summary_with_llm(summary)

        row["EmbeddingText"] = embedding_text

    except Exception as e:
        logging.error(f"Failed → {pdf_path}: {e}")
        row["Summary"] = "NA"
        row["EmbeddingText"] = "NA"

    return row

# ============================================================

def is_circular(subcategory: str) -> bool:
    if not isinstance(subcategory, str):
        return False

    # Normalize
    sub = subcategory.strip().lower()

    #  Explicitly exclude Master Circulars
    if "master circular" in sub:
        return False

    # Normalize spacing around hyphens (e.g., "circular - bse" → "circular-bse")
    sub = re.sub(r'\s*-\s*', '-', sub)

    #  Allow only standard circular variants
    return sub in {
        "circular",
        "circulars",
        "circular-bse",
        "circular-nse"
    }

def process_circular_pdf(row: pd.Series):
    pdf_path = Path(row["Path"])

    try:
        text = extract_pdf_text(pdf_path)

        # Circulars do NOT need regulation-style filtering
        core_text = text[:12000]

        summary = llm.invoke(
            CIRCULARS_PROMPT.format(text=core_text)
        ).strip()

        summary = summary.strip().strip('"')

        # row["Summary"] = summary or "NA"
        row["Summary"] = clean_summary_with_llm(summary)

        row["EmbeddingText"] = core_text

    except Exception as e:
        logging.error(f"Failed → {pdf_path}: {e}")
        row["Summary"] = "NA"
        row["EmbeddingText"] = "NA"

    return row

# ============================================================

def process_press_release_pdf(row: pd.Series):
    pdf_path = Path(row["Path"])

    try:
        text = extract_pdf_text(pdf_path)
        text = re.sub(r'\s+', ' ', text).strip()
        core_text = text[:4000]

        summary = llm.invoke(
            PRESS_RELEASE_PROMPT.format(text=core_text)
        ).strip()

        # row["Summary"] = summary or "NA"
        row["Summary"] = clean_summary_with_llm(summary)

        row["EmbeddingText"] = core_text

    except Exception as e:
        logging.error(f"Failed → {pdf_path}: {e}")
        row["Summary"] = "NA"
        row["EmbeddingText"] = "NA"

    return row

# ============================================================

def clean_consultation_title(title: str) -> str:
    if not isinstance(title, str):
        return title

    # Remove trailing "Click here to provide your comments" (case-insensitive)
    cleaned = re.sub(
        r'\s*click here to provide your comments\s*$',
        '',
        title,
        flags=re.IGNORECASE
    )

    return cleaned.strip()

def process_consultation_paper_pdf(row: pd.Series):
    pdf_path = Path(row["Path"])

    try:
        text = extract_pdf_text(pdf_path)

        # No regulation-style filtering required
        core_text = text[:12000]

        summary = llm.invoke(
            CONSULTATION_PAPER_PROMPT.format(text=core_text)
        ).strip()

        # row["Summary"] = summary or "NA"
        row["Summary"] = clean_summary_with_llm(summary)

        row["EmbeddingText"] = core_text

    except Exception as e:
        logging.error(f"Failed → {pdf_path}: {e}")
        row["Summary"] = "NA"
        row["EmbeddingText"] = "NA"

    return row

# ============================================================

def is_master_circular(subcategory: str) -> bool:
    if not isinstance(subcategory, str):
        return False
    return "master circular" in subcategory.lower()


def extract_master_circular_core(text: str) -> str:
    lines = text.splitlines()
    keep = []

    for line in lines:
        clean = line.strip()

        if not clean:
            continue

        if re.search(r'table of contents|contents|index', clean, re.IGNORECASE):
            break

        if len(clean) > 30:
            keep.append(clean)

        if len(keep) >= 25:
            break

    return "\n".join(keep)

def process_master_circular_pdf(row: pd.Series):
    pdf_path = Path(row["Path"])

    try:
        text = extract_pdf_text(pdf_path)

        #  Only intro part, not full document
        core_text = extract_master_circular_core(text)

        summary = llm.invoke(
            MASTER_CIRCULAR_PROMPT.format(text=core_text)
        ).strip()
        summary = re.sub(r'https?://\S+', '', summary)

        # row["Summary"] = summary or "NA"
        row["Summary"] = clean_summary_with_llm(summary) or "NA"

        row["EmbeddingText"] = core_text

    except Exception as e:
        logging.error(f"Failed → {pdf_path}: {e}")
        row["Summary"] = "NA"
        row["EmbeddingText"] = "NA"

    return row

# ============================================================

def is_notification(subcategory: str) -> bool:
    if not isinstance(subcategory, str):
        return False
    return subcategory.strip().lower() == "notifications"

def process_notification_pdf(row: pd.Series):
    pdf_path = Path(row["Path"])

    try:
        text = extract_pdf_text(pdf_path)

        # Notifications are short → no regulation-style filtering
        core_text = text[:8000]

        summary = llm.invoke(
            NOTIFICATIONS_PROMPT.format(text=core_text)
        ).strip()

        row["Summary"] = clean_summary_with_llm(summary)
        row["EmbeddingText"] = core_text

    except Exception as e:
        logging.error(f"Failed → {pdf_path}: {e}")
        row["Summary"] = "NA"
        row["EmbeddingText"] = "NA"

    return row

# ============================================================

def process_row_by_domain(row: pd.Series):
    sub = row["SubCategory"]

    if not isinstance(sub, str):
        logging.warning("SubCategory is missing or invalid")
        return None

    sub_clean = sub.strip().lower()

    #  Master Circular FIRST (important)
    if is_master_circular(sub_clean):
        return process_master_circular_pdf(row)

    elif sub_clean == "regulations":
        return process_regulation_pdf(row)

    elif is_circular(sub_clean):
        return process_circular_pdf(row)

    elif "press release" in sub_clean:
        return process_press_release_pdf(row)

    elif sub_clean == "consultation paper":
        return process_consultation_paper_pdf(row)
    
    elif is_notification(sub_clean):
        return process_notification_pdf(row)

    else:
        logging.info(f"Skipping SubCategory → {sub}")
        return None

# ============================================================
# MAIN
# ============================================================

def main(excel_file: str):
    df = pd.read_excel(excel_file)

    required = ["Verticals", "SubCategory", "Path"]
    for col in required:
        if col not in df.columns:
            raise ValueError(f"Missing column: {col}")

    # 🔹 Clean Consultation Paper titles
    if "Title" in df.columns:
        df["Title"] = df.apply(
            lambda r: clean_consultation_title(r["Title"])
            if isinstance(r["SubCategory"], str)
            and r["SubCategory"].strip().lower() == "consultation paper"
            else r["Title"],
            axis=1
        )

    logging.info(f"Processing {len(df)} PDFs across all subcategories")

    start = time.time()

    for idx, row in df.iterrows():
        logging.info(f"[{idx+1}/{len(df)}] {row['Path']}")
        processed = process_row_by_domain(row)
        if processed is None:
            continue    
        update_excel(processed)

    logging.info(f"Completed in {time.time() - start:.2f}s")
    
    # -------------------------------------------------
    # EXACT PLACE FOR MINIO UPLOAD 
    # -------------------------------------------------

    try:
        minio = MinIOClient()

        week_folder_name = WEEK_FOLDER.name
        minio_prefix = f"weekly_outputs/{week_folder_name}/"

        #  HARD RESET week folder
        minio.delete_prefix(minio_prefix)

        #  Upload ONLY Excel files created in THIS run
        for excel_name in CREATED_EXCELS:
            local_excel = WEEK_FOLDER / excel_name
            object_path = f"{minio_prefix}{excel_name}"  # ✅ NO DOUBLE SLASH

            minio.upload_file(
                local_path=str(local_excel),
                object_path=object_path
            )

        logging.info(
            f"Uploaded weekly Excel files to MinIO → "
            f"bucket={minio.bucket}, prefix={minio_prefix}"
        )

    except Exception as e:
        logging.error(f"MinIO upload failed: {e}")

# ============================================================
# ENTRY
# ============================================================

if __name__ == "__main__":
    excel = DATA_DIR / "Searching_agent_output.xlsx"
    if not excel.exists():
        raise FileNotFoundError("Searching_agent_output.xlsx not found")

    main(excel)