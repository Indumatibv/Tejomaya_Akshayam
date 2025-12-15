#Using MiniLM embeddings(all-MiniLM-L6-v2) 
# Making embeddings of chunk summaries to improve final summary relevance and reduce hallucinations.
#!/usr/bin/env python
# agents/parsing_agent.py

import os
import re
import logging
from pathlib import Path
import pandas as pd
import nltk
from dotenv import load_dotenv
from unstructured.partition.pdf import partition_pdf
from langchain.text_splitter import RecursiveCharacterTextSplitter
from langchain.llms import Ollama
from langchain.prompts import PromptTemplate
from openpyxl import load_workbook, Workbook
import torch
import warnings
import time

# ---- NEW IMPORTS FOR FAISS + MINILM ----
from langchain_community.vectorstores import FAISS
from langchain_community.embeddings import HuggingFaceEmbeddings

# Load MiniLM embedding model (local / free)
emb_model = HuggingFaceEmbeddings(model_name="sentence-transformers/all-MiniLM-L6-v2")


# ---------------------- Logging ----------------------
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[logging.StreamHandler()]
)

warnings.filterwarnings("ignore", message="No languages specified, defaulting to English.")
nltk.download('punkt', quiet=True)
nltk.download('averaged_perceptron_tagger', quiet=True)
load_dotenv()

# ---------------------- GPU Detection (same as original) ----------------------
device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
os.environ["OLLAMA_USE_GPU"] = "1" if device.type == "cuda" else "0"
logging.info(f"💪 Using device: {device}")
logging.info("Ollama GPU status: " + ("Enabled" if os.environ.get("OLLAMA_USE_GPU") == "1" else "Disabled"))

# ---------------------- Initialize LLM ----------------------
llm = Ollama(model="mistral:latest")

# ---------------------- Prompts ----------------------

chunk_prompt = PromptTemplate(
    template="""
    You are analyzing a regulatory or financial document.
    Summarize the following section in a **concise and factual** manner (2 sentences max).

    Focus only on:
    - The key regulatory or policy action (amendment, relaxation, proposal, or update)
    - Who or what it affects
    - Any critical dates or deadlines mentioned
    
    Avoid:
    - Legal citations, section numbers, or procedural details
    - Repetition of headers, page numbers, or contact information
    - Explanations, reasoning, or generic filler text

    Chunk Text:
    {text}

    Short factual summary:
    """,
    input_variables=["text"]
)

# Final prompt converted to bullet-style output (•). No double quotes required.
final_prompt = PromptTemplate(
    template="""
    You are a senior communications analyst summarizing official circulars or notices for a mixed audience 
    that may include company stakeholders, investors, and general readers. 
    Produce a concise bullet-point summary (use filled-dot bullets '•') that helps readers instantly understand the circular’s purpose, 
    key changes, applicability, and practical impact — without needing to read the original document.

    STRICT GUIDELINES:
    - Use plain, professional, and neutral language.
    - Output exactly 4–6 bullet points where applicable; fewer if the source contains fewer clear facts.
    - The summary must be fully self-contained and understandable on its own.
    - Rephrase any technical or legal content into simple, direct meaning.
    - Do NOT invent any detail that is not explicitly present in the provided text.
    - Avoid legal citations, clause numbers, regulation names, annexures, or procedural instructions.
    - Avoid filler, speculation, or procedural instructions (e.g., how to file or whom to email).

    FOCUS ON:
        • What the circular is about and why it was issued  
        • The key changes, relaxations, thresholds, exemptions, or requirements introduced  
        • Who it applies to (e.g., listed companies, shareholders, investors)  
        • Important dates or conditions that define applicability  
        • How the update affects compliance steps or business practices  

    AVOID COMPLETELY:
        • Mentions of clause numbers, paragraph numbers, annexures, or regulation names 
        • Phrases that imply further reading is needed (e.g., “as specified”, “as per the circular”)
        • Legal jargon, disclaimers, or procedural references like emails, filings, or submission methods  

    Text:
    {text}

    Final bullet summary (use '•' for each bullet; do NOT wrap output in quotes; no numbered lists):
    """,
    input_variables=["text"]
)


# ---------------------- Folders ----------------------
BASE_DIR = Path(__file__).resolve().parent.parent  # Tejomaya_ETL_pipeline/
DATA_DIR = BASE_DIR / "data"
OUTPUT_EXCEL_DIR = DATA_DIR / "output_excels"
OUTPUT_EXCEL_DIR.mkdir(parents=True, exist_ok=True)

from datetime import datetime, timedelta

# ---------------------- Weekly Folder Creation ----------------------
def get_week_folder():
    import json

    # Read week range from searching agent
    week_json = DATA_DIR / "week_range.json"
    if not week_json.exists():
        raise ValueError("week_range.json not found! Run searching_agent first.")

    with open(week_json, "r") as f:
        week_info = json.load(f)

    week_start = datetime.strptime(week_info["week_start"], "%Y-%m-%d")
    week_end = datetime.strptime(week_info["week_end"], "%Y-%m-%d")

    folder_name = f"{week_start.strftime('%Y-%m-%d')}_to_{week_end.strftime('%Y-%m-%d')}"
    week_folder = OUTPUT_EXCEL_DIR / folder_name

    # Replace folder ONLY if same week
    if week_folder.exists():
        import shutil
        shutil.rmtree(week_folder)

    week_folder.mkdir(parents=True, exist_ok=True)
    return week_folder


WEEK_FOLDER = get_week_folder()
logging.info(f"Weekly output folder → {WEEK_FOLDER}")

# ---------------------- PDF Cleaning ----------------------
def clean_pdf_text(text: str) -> str:
    pages = re.split(r'\f+', text)
    cleaned_pages = []

    for page in pages:
        lines = page.splitlines()
        if len(lines) > 4:
            lines = lines[2:-2]

        page_text = "\n".join(lines)
        page_text = re.sub(r'Page\s*\d+\s*(of\s*\d+)?', '', page_text, flags=re.IGNORECASE)
        page_text = re.sub(r'(Securities and Exchange Board of India|Consultation Paper|Master Circular)', '', page_text, flags=re.IGNORECASE)

        cleaned_pages.append(page_text.strip())

    cleaned = "\n".join(cleaned_pages)
    cleaned = re.sub(r'\n{2,}', '\n', cleaned)
    cleaned = re.sub(r'\s{2,}', ' ', cleaned)
    return cleaned.strip()

# ---------------------- Language Detection ----------------------
def detect_language(text: str) -> str:
    devanagari = len(re.findall(r'[\u0900-\u097F]', text))
    latin = len(re.findall(r'[A-Za-z]', text))
    total = devanagari + latin
    if total == 0:
        return "unknown"

    if devanagari / total > 0.3 and latin / total > 0.3:
        return "mixed"
    if devanagari / total > 0.3:
        return "hindi"
    if latin / total > 0.3:
        return "english"
    return "unknown"

# ---------------------- Filter English ----------------------
def filter_english_text(text: str) -> str:
    return "\n".join(
        line for line in text.split("\n")
        if re.search(r'[A-Za-z]', line) and not re.search(r'[\u0900-\u097F]', line)
    )

# ---------------------- Index Extraction ----------------------
def extract_indexing_from_first_page(pdf_path: str) -> str:
    try:
        first_page_data = partition_pdf(
            filename=str(pdf_path),
            strategy="fast",
            include_page_breaks=False,
            starting_page_number=1,
            max_pages=1
        )
        text = "\n".join(str(el) for el in first_page_data if el)

        pattern = r"""
        \b(
            REGD\.?\s*No\.?\s*[A-Z.\-\s]*\d+(?:/\d+)* |
            (?:CG|DL|MH|HR|UP|GJ|TN|RJ|KL|KA|WB|PB|CH|UK|AS|OR|BR|AP|TS|HP|GA|JK|NL|MN|TR|SK|AR)-[A-Z]{2}-E-\d{8}-\d+ |
            No\.?\s*[A-Z/.\-]*\d+(?:/\d+)* |
            SEBI/[A-Z]{2,}/\d{2}/\d{2} |
            S\.O\.\s*\d+\(E\) |
            IBBI/\d{4}-\d{2}/GN/REG\d+
        )\b
        """
        matches = re.findall(pattern, text, flags=re.VERBOSE)
        return ", ".join(matches) if matches else "NA"

    except Exception as e:
        logging.warning(f"Indexing extraction failed for {pdf_path}: {e}")
        return "NA"

# ---------------------- Extract PDF ----------------------
def extract_pdf_text(pdf_path: str):
    pdf_path = Path(pdf_path).resolve()
    indexing = extract_indexing_from_first_page(pdf_path)

    raw = partition_pdf(filename=str(pdf_path), strategy="fast", include_page_breaks=False)
    extracted = "\n".join(str(el) for el in raw if el).strip()

    if not extracted:
        logging.info(f"Using hi_res OCR for: {pdf_path}")
        raw = partition_pdf(filename=str(pdf_path), strategy="hi_res", extract_images_in_pdf=True)
        extracted = "\n".join(str(el) for el in raw if el).strip()

    cleaned = clean_pdf_text(extracted)
    lang = detect_language(cleaned) if cleaned else "unknown"

    logging.info(f"Lang: {lang} | Index: {indexing}")
    return cleaned, lang, indexing

# ---------------------- UPDATED Summary Generation (FAISS Integrated, fixed prompt formatting) ----------------------
def generate_summary(extracted_text: str, max_tokens: int = 1000):
    try:
        splitter = RecursiveCharacterTextSplitter(chunk_size=4000, chunk_overlap=500)
        docs = splitter.create_documents([extracted_text])

        chunk_summaries = []

        # ---- Step 1: Generate chunk summaries ----
        for doc in docs:
            summary = llm.invoke(chunk_prompt.format(text=doc.page_content)).strip()
            if summary:
                chunk_summaries.append(summary)

        if not chunk_summaries:
            return {"embedding_text": "NA", "client_summary": "NA"}

        # ---- Step 2: Create temporary FAISS index ----
        faiss_index = FAISS.from_texts(chunk_summaries, emb_model)

        # Use a small top-k to reduce hallucination risk
        top_k = min(5, len(chunk_summaries))
        retrieved_docs = faiss_index.similarity_search("overall summary", k=top_k)

        # Concatenate retrieved chunk summaries in ranked order
        retrieved_text = "\n".join([d.page_content for d in retrieved_docs])

        # Truncate conservatively for the final prompt
        truncated = retrieved_text[:max_tokens * 4]

        # ---- Step 3: Generate FINAL summary using correct PromptTemplate formatting ----
        final_query = final_prompt.format_prompt(text=truncated).to_string()
        final = llm.invoke(final_query).strip()

        # ---- Step 4: Delete FAISS index (temporary only) ----
        try:
            del faiss_index
        except Exception:
            pass

        return {
            "embedding_text": retrieved_text,
            "client_summary": final or "NA"
        }

    except Exception as e:
        logging.error(f"Summarization failed: {e}")
        return {"embedding_text": "NA", "client_summary": "NA"}

# ---------------------- Excel Update ----------------------
def update_excel(row: pd.Series):
    vertical = row["Verticals"]
    sub = row["SubCategory"]

    excel_path = WEEK_FOLDER / f"{vertical}.xlsx"

    if excel_path.exists():
        wb = load_workbook(excel_path)
    else:
        wb = Workbook()
        wb.remove(wb.active)

    if sub in wb.sheetnames:
        ws = wb[sub]
    else:
        ws = wb.create_sheet(title=sub)
        ws.append(list(row.index))

    ws.append([row.get(col, "NA") or "NA" for col in row.index])
    wb.save(excel_path)
    wb.close()

    logging.info(f"Updated Excel → {excel_path} ({sub})")

# ---------------------- Process One PDF ----------------------
def process_single_pdf(row: pd.Series):
    pdf_path = row["Path"]
    try:
        extracted, lang, indexing = extract_pdf_text(pdf_path)
        row["Indexing"] = indexing

        if extracted:
            if lang in ["english", "mixed"]:
                english = filter_english_text(extracted)
                summaries = generate_summary(english)
                row["Summary"] = summaries["client_summary"]
                row["EmbeddingText"] = summaries["embedding_text"]
            elif lang == "hindi":
                row["Summary"], row["EmbeddingText"] = "FULL_HINDI", "NA"
            else:
                row["Summary"], row["EmbeddingText"] = "NA", "NA"
        else:
            row["Summary"], row["EmbeddingText"] = "NA", "NA"

    except Exception as e:
        logging.error(f"Error processing PDF {pdf_path}: {e}")
        row["Summary"], row["EmbeddingText"], row["Indexing"] = "NA", "NA", "NA"

    return row

# ---------------------- Main Pipeline ----------------------
def main(excel_file: str):
    excel_file = Path(excel_file)
    if not excel_file.exists():
        raise FileNotFoundError(f"Input Excel not found: {excel_file}")

    df = pd.read_excel(excel_file)

    required = ["Verticals", "SubCategory", "Path"]
    for col in required:
        if col not in df.columns:
            raise ValueError(f"Excel missing required column: {col}")

    df["Indexing"] = df.get("Indexing", "")
    df["Summary"] = df.get("Summary", "")
    df["EmbeddingText"] = df.get("EmbeddingText", "")

    total = len(df)
    logging.info(f"Starting PDF processing: {total} files")

    start = time.time()

    for idx, row in df.iterrows():
        logging.info(f"Processing PDF {idx+1}/{total}: {row['Path']}")
        processed = process_single_pdf(row)
        update_excel(processed)

    logging.info(f"All PDFs processed in {time.time() - start:.2f}s")

# ---------------------- Entry ----------------------
if __name__ == "__main__":
    import sys

    default_excel = DATA_DIR / "weekly_sebi_downloads.xlsx"

    if len(sys.argv) >= 2:
        excel_file = Path(sys.argv[1])
        logging.info(f"Using Excel from CLI → {excel_file}")
    elif default_excel.exists():
        excel_file = default_excel
        logging.info(f"Auto-detected Excel → {excel_file}")
    else:
        logging.error("No Excel found! Exiting.")
        sys.exit(1)

    try:
        main(excel_file)
    except Exception as e:
        logging.error(f"Fatal error: {e}")
